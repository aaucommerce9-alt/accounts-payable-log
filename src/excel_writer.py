"""
Write enriched transaction rows into the Excel workbook.

- Appends only new rows to the Transactions tab (dedup on key).
- Refreshes the Monthly Summary tab from scratch each run.
- Creates the workbook and tabs if they don't yet exist.
"""

import logging
from collections import defaultdict
from datetime import date as date_type
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _col_map(cfg: dict) -> dict[str, int]:
    """Return {field_name: column_index} from config."""
    return cfg["excel"]["columns"]


def _row_key(row: dict) -> str:
    return f"{row['order_id']}|{row['settlement_id']}|{row['type']}|{row['sku']}|{row['date']}"


def _ensure_workbook(wb_path: str, cfg: dict) -> openpyxl.Workbook:
    p = Path(wb_path)
    if p.exists():
        return openpyxl.load_workbook(wb_path)
    logger.info("Workbook not found — creating new: %s", wb_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default Sheet
    _ensure_tabs(wb, cfg)
    wb.save(wb_path)
    return wb


def _ensure_tabs(wb: openpyxl.Workbook, cfg: dict):
    txn_tab = cfg["excel"]["transactions_tab"]
    sum_tab = cfg["excel"]["monthly_summary_tab"]
    if txn_tab not in wb.sheetnames:
        ws = wb.create_sheet(txn_tab)
        _write_txn_header(ws, cfg)
    if sum_tab not in wb.sheetnames:
        wb.create_sheet(sum_tab)


def _write_txn_header(ws, cfg: dict):
    cols = _col_map(cfg)
    header_map = {v: k.replace("_", " ").title() for k, v in cols.items()}
    for col_idx, label in sorted(header_map.items()):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")


# ---------------------------------------------------------------------------
# Public interface
# ---------------------------------------------------------------------------

def load_existing_keys(wb_path: str, cfg: dict) -> set[str]:
    """Return set of row keys already in the Transactions tab."""
    keys: set[str] = set()
    p = Path(wb_path)
    if not p.exists():
        return keys
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    tab = cfg["excel"]["transactions_tab"]
    if tab not in wb.sheetnames:
        wb.close()
        return keys
    ws    = wb[tab]
    cols  = _col_map(cfg)
    c_oid = cols["order_id"]
    c_sid = cols["settlement_id"]
    c_typ = cols["type"]
    c_sku = cols["sku"]
    c_dt  = cols["date"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        oid = row[c_oid - 1] or ""
        sid = row[c_sid - 1] or ""
        typ = row[c_typ - 1] or ""
        sku = row[c_sku - 1] or ""
        dt  = row[c_dt  - 1] or ""
        keys.add(f"{oid}|{sid}|{typ}|{sku}|{dt}")
    wb.close()
    return keys


def append_transactions(rows: list[dict], wb_path: str, cfg: dict) -> int:
    """Append new rows; return count of rows written."""
    if not rows:
        return 0

    existing_keys = load_existing_keys(wb_path, cfg)
    new_rows = [r for r in rows if _row_key(r) not in existing_keys]
    if not new_rows:
        logger.info("No new rows to write (all %d rows already present).", len(rows))
        return 0

    wb  = _ensure_workbook(wb_path, cfg)
    _ensure_tabs(wb, cfg)
    ws  = wb[cfg["excel"]["transactions_tab"]]
    cols = _col_map(cfg)

    # Find next empty row
    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(1, 1).value is None:
        # Brand-new sheet with no header yet
        _write_txn_header(ws, cfg)
        next_row = 2

    field_order = {field: idx for field, idx in cols.items()}

    for row in new_rows:
        for field, col_idx in field_order.items():
            val = row.get(field, "")
            if isinstance(val, float) and field.endswith("_pct"):
                # Store percentages as decimals for Excel formatting
                val = val / 100
            ws.cell(row=next_row, column=col_idx, value=val)
        next_row += 1

    wb.save(wb_path)
    logger.info("Wrote %d new rows to Transactions tab.", len(new_rows))
    return len(new_rows)


def refresh_monthly_summary(wb_path: str, cfg: dict):
    """Rebuild the Monthly Summary tab from Transactions data."""
    wb = _ensure_workbook(wb_path, cfg)
    _ensure_tabs(wb, cfg)

    txn_ws  = wb[cfg["excel"]["transactions_tab"]]
    sum_ws  = wb[cfg["excel"]["monthly_summary_tab"]]
    cols    = _col_map(cfg)

    # Aggregate by year-month
    # Keys: "YYYY-MM", values: dict of sums
    monthly: dict[str, dict] = defaultdict(lambda: defaultdict(float))

    numeric_fields = [
        "gross_sales", "promotions", "referral_fee", "fba_fee",
        "other_fees", "refunds", "sales_tax", "net_proceeds",
        "cogs", "profit_share", "net_profit",
    ]

    c_date = cols["date"]

    for row in txn_ws.iter_rows(min_row=2, values_only=True):
        dt = row[c_date - 1]
        if not dt:
            continue
        if isinstance(dt, str):
            try:
                from datetime import date as _d
                dt = _d.fromisoformat(dt)
            except Exception:
                continue
        month_key = dt.strftime("%Y-%m")
        for field in numeric_fields:
            val = row[cols[field] - 1] or 0
            monthly[month_key][field] += float(val)

    # Clear existing summary content (keep tab)
    sum_ws.delete_rows(1, sum_ws.max_row or 1)

    # Header
    summary_cols = ["Month"] + [f.replace("_", " ").title() for f in numeric_fields]
    for i, h in enumerate(summary_cols, 1):
        cell = sum_ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    # Data rows, sorted chronologically
    for r_idx, month_key in enumerate(sorted(monthly.keys()), 2):
        sum_ws.cell(row=r_idx, column=1, value=month_key)
        for c_idx, field in enumerate(numeric_fields, 2):
            sum_ws.cell(row=r_idx, column=c_idx, value=round(monthly[month_key][field], 2))

    wb.save(wb_path)
    logger.info("Monthly Summary refreshed (%d months).", len(monthly))
