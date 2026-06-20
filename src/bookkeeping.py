"""
Apply COGS lookup, profit-share accrual, and derived metrics to each row.
"""

import logging
from datetime import date as date_type
from src.config_loader import profit_share_rate

logger = logging.getLogger(__name__)


def load_cost_master(workbook_path: str, cfg: dict) -> dict[str, dict]:
    """
    Returns {sku: {"cogs": float, "supplier": str}} from the Cost Master sheet.
    Warns if the sheet or columns are missing.
    """
    import openpyxl
    cost_map: dict[str, dict] = {}
    tab_name = cfg["paths"]["cost_master_tab"]
    sku_col  = cfg["cost_master"]["sku_col"]
    cogs_col = cfg["cost_master"]["cogs_col"]
    sup_col  = cfg["cost_master"]["supplier_col"]

    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except FileNotFoundError:
        logger.warning("Workbook not found at '%s'; COGS will be 0 for all SKUs.", workbook_path)
        return cost_map

    if tab_name not in wb.sheetnames:
        logger.warning("Cost Master tab '%s' not found; COGS will be 0 for all SKUs.", tab_name)
        wb.close()
        return cost_map

    ws = wb[tab_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        sku_idx  = headers.index(sku_col)
        cogs_idx = headers.index(cogs_col)
        sup_idx  = headers.index(sup_col)
    except ValueError as e:
        logger.warning("Cost Master column missing: %s", e)
        wb.close()
        return cost_map

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[sku_idx]).strip() if row[sku_idx] else None
        if not sku:
            continue
        cost_map[sku] = {
            "cogs":     float(row[cogs_idx] or 0),
            "supplier": str(row[sup_idx] or "").strip(),
        }

    wb.close()
    logger.info("Loaded COGS for %d SKUs from Cost Master.", len(cost_map))
    return cost_map


def apply_calculations(rows: list[dict], cost_map: dict[str, dict], cfg: dict) -> list[dict]:
    """Enrich each row with COGS, profit-share, net_profit, margin, ROI."""
    ps_base = cfg["profit_share"]["base"]
    missing_skus: set[str] = set()

    for row in rows:
        sku  = row.get("sku", "")
        info = cost_map.get(sku, {})
        cogs     = info.get("cogs", 0.0)
        supplier = info.get("supplier", "")

        if sku and sku not in cost_map:
            missing_skus.add(sku)

        row["cogs"]     = cogs
        row["supplier"] = supplier

        txn_date = row.get("date")
        if isinstance(txn_date, str):
            from datetime import date as _d
            txn_date = _d.fromisoformat(txn_date) if txn_date else None

        ps_rate = profit_share_rate(cfg, txn_date or date_type.today(), supplier)

        # Determine profit-share base
        gross = row.get("gross_sales", 0.0)
        net   = row.get("net_proceeds", 0.0)
        promos = abs(row.get("promotions", 0.0))
        refunds = abs(row.get("refunds", 0.0))

        if ps_base == "gross_profit":
            base_amt = gross - cogs
        elif ps_base == "net_sales":
            base_amt = gross - promos - refunds
        else:  # net_profit (default)
            base_amt = net - cogs

        profit_share = round(max(base_amt, 0.0) * ps_rate, 2)
        net_profit   = round(net - cogs - profit_share, 2)
        margin_pct   = round(net_profit / gross * 100, 4) if gross else 0.0
        roi_pct      = round(net_profit / cogs * 100, 4)  if cogs  else 0.0

        row["profit_share"] = profit_share
        row["net_profit"]   = net_profit
        row["margin_pct"]   = margin_pct
        row["roi_pct"]      = roi_pct

    if missing_skus:
        logger.warning(
            "COGS not found for %d SKU(s): %s — set to 0. "
            "Add them to your Cost Master sheet.",
            len(missing_skus), ", ".join(sorted(missing_skus))
        )

    return rows
